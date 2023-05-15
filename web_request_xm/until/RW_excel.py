'''读取excel文件类和方法的封装
场景1： 把接口测试用例的数据 参数化，写到一个excel表格
     1.读取所有excel的数据,代表所有的用例，
       组装成什么的数据类型
        目前只支持第一个sheet工作表，
        需求作业：包括excel所有的sheet工作表的用例
        [
            （一个用例），
            （二个用例），
            （'admin','11234343'）
        ]
        [
            [一个用例]，
            [二个用例]，
            [三个用例]
        ]
        [
            {
                "username":"admin"
                "password":"123456"
            },
            {},
            {}，
        ]
     2.ddt数据驱动,把excel表格每一行的数据转换成用例
'''
import xlrd
import openpyxl

class ExcelWR():
    def __init__(self,filepath=r'D:\1967668484\git_rc\jk_zdh\10qf_frame_code\qf_frame\config\jk-sj.xlsx'):
        self.filepath=filepath
        self.excel=xlrd.open_workbook(self.filepath)
        self.book=openpyxl.load_workbook(self.filepath)

    def readEx(self,runCases=None):
        '''
        1.指定运行某一个用例   cases=(1,[1])  元祖第一个元素，代表的sheet工作表
        2.根据用例的编号 	   cases=(1,[1,3,6,8,10])
        3.依次运行批量的用例   cases=(1,(2,4))
        :param cases:
        :return:
        '''
        datas = []  # 定义一个空列表，
        # 1.表示需要指定用例运行
        if runCases:
            sheet = self.excel.sheet_by_index(runCases[0]-1)
            #  #cases=(1,[1])
            if isinstance(runCases[1],list): #runCases[1]=[2,4,6]
                for i in runCases[1]: #i=2,4,6
                    value = sheet.row_values(i) #第二行的数据
                    datas.append(value)
            elif isinstance(runCases[1],tuple): #runCases[1]=(2,5) ，第2条-第五条，包括第5条
                 for i in range(runCases[1][0],runCases[1][1]+1): #i=2,3,4
                     value = sheet.row_values(i) #i=4,第五行的数据
                     datas.append(value)
        else:
            '''2.默认所有的用例'''
            sheets = self.excel.sheet_names()  # 获取excel表格所有的sheet工作表,返回的是列表
            print('接口自动化用例文件有{}个sheet工作表'.format(len(sheets)))
            for s in range(len(sheets) - 1):  # s=0,1
                sheet = self.excel.sheet_by_index(s)  # 一次获取每个工作表
                rows, cols = sheet.nrows, sheet.ncols
                for i in range(1, rows):  # rows=10,取1,2,3,4。。。。9
                    value = sheet.row_values(i)
                    datas.append(value)
        return datas

    def write_tiqu(self,name,value):
        '''把提取的变量写入到最后一个sheet工作表'''
        sheet=self.book['公共变量']
        # sheet.cell(2,1).value=name
        # sheet.cell(2,2).value=value
        '''公共变量里面的变量名称肯定是唯一，
        所以写入的时候，如果存在就更新
        循环 excel的行数
        '''
        rows=sheet.max_row #工作表的行数,3
        for i in range(2,rows+1): #i=2,3
            print(sheet.cell(i,1).value)
            #表示已经存在，存在就更新  token_tiqu
            #能够走进if的代码，说明存在相同的，更新值之后，退出函数
            if sheet.cell(i,1).value==name:
                sheet.cell(i,2).value=value
                self.book.save(self.filepath)
                return
        #循环结束没有走到if里面去，说明变量文件不存在重复的名称
        sheet.append([name,value])
        self.book.save(self.filepath)

    def read_tiqu(self,name):
        name=name+'_tiqu'
        sheet=self.excel.sheet_by_name('公共变量')
        for i in range(1,sheet.nrows): #i=1,2,3
            value=sheet.row_values(i)[0] #获取每一行第一列的值
            if value==name:
                return sheet.row_values(i)[1]

'''
	需求：
    现在运行代码，默认是依次运行excel中所有的用例
    1.指定运行某一个用例   (1,[1])
    2.根据用例的编号 	   (1,[1,3,6,8,10])
    3.依次运行批量的用例   (1,(2,8))

'''
def get_case(cases=None):
    '''获取用例的所有的数据'''
    return  ExcelWR().readEx(cases)

def write_bl(name,value):
    '''写入变量'''
    return  ExcelWR().write_tiqu(name,value)

def read_bl(name):
    '''读取变量'''
    return  ExcelWR().read_tiqu(name)



if __name__ == '__main__':
    # E=ExcelWR(filepath=r'../config/testcase.xlsx')
    # #caseDatas=E.readEx()
    # #E.write_tiqu('token_tiqu','xxxxxx'),
    # print(E.read_tiqu('city'))
    print(get_case((1,(2,6))))





