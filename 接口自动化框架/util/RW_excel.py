'''读取excel文件类和方法的封装'''
import xlrd
import openpyxl


class ExcelWR():

    def __init__(self, filePath):
        self.filePath = filePath
        self.excel = xlrd.open_workbook(self.filePath)  # 1、打开
        self.book = openpyxl.load_workbook(self.filePath)

    def readEx(self):
        sheets = self.excel.sheet_names()  # 2、 获取excel表格所有的sheet工作表，返回的是列表
        # print(sheets)
        datas = []
        for s in range(len(sheets) - 1):
            # print(s)
            sheet = self.excel.sheet_by_index(s)  # 第一个工作表
            rows, cols = sheet.nrows, sheet.ncols
            # print(rows,cols)

            for i in range(1, rows):  # 从第二行开始，2,3,4,5,6。rows为6。 arr[5]刚好是最后一行
                value = sheet.row_values(i)
                # print(value)
                datas.append(value)
                # print(datas)
        return datas

    '''把提取的变量写入最后一个sheet工作表'''

    def write_tiqu(self, name, value):
        sheet = self.book['公共变量']  # 指定表格
        # sheet.cell(2,1).value=name        #第二行第一个值
        # sheet.cell(2,2).value=value       #第二行第二个值
        rows = sheet.max_row
        for i in range(2, rows + 1):
            print(sheet.cell(i, 1).value)
            if sheet.cell(i, 1).value == name:  # if重复值
                sheet.cell(i, 2).value = value  # 循环结束没有走到if里面去，说明变量文件不存在重复的名称
                self.book.save(self.filePath)
                return
        sheet.append([name, value])  # append方法
        self.book.save(self.filePath)  # 保存

        # 读取变量

    def read_tiqu(self, name):
        name = name + '_tiqu'
        sheet = self.excel.sheet_by_name('公共变量')
        for i in range(1, sheet.nrows):  # i=1,2,3
            value = sheet.row_values(i)[0]  # 获取每一行每一列的值
            if value == name:
                return sheet.row_values(i)[1]


def get_case(self):
    return ExcelWR().readEx()


def write_bl():
    return ExcelWR().write_tiqu(name,value)

def read_bl():
    return ExcelWR().read_tiqu(name)


if __name__ == '__main__':
    E = ExcelWR(filePath=r'D:\1967668484\git-资料测试总理-已提交\接口自动化\接口自动化框架\config\jk-sj.xlsx')  # 路径
    # caseDatas = E.readEx()
    # E.write_tiqu('token_tiqu','qingfengtest2')
    print(E.read_tiqu('token'))
